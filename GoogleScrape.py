# Loading Libraries
import warnings
from openpyxl import load_workbook
import pandas as pd
import glob
import os
import time
from googlesearch import search
import pandas as pd
import numpy as np
import re
import datetime
from bs4 import BeautifulSoup
from newspaper import Article
import trafilatura
import requests
import nltk
import gensim
from gensim import corpora
from gensim.parsing.preprocessing import strip_non_alphanum
from gensim.parsing.preprocessing import strip_multiple_whitespaces
from gensim.parsing.preprocessing import remove_stopwords
import xlsxwriter
import matplotlib.pyplot as plt
import string
import pprint
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk import (
    TrigramAssocMeasures,
    TrigramCollocationFinder,
    BigramAssocMeasures,
    BigramCollocationFinder,
)
from operator import itemgetter
import spacy
from spacy.lang.en import English
import en_core_web_sm
from textblob import TextBlob
from wordcloud import WordCloud
from collections import Counter
import pyLDAvis
import pyLDAvis.gensim_models
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import networkx as nx
import textacy
from spacy.matcher import Matcher
from spacy.tokens import Span
from nltk.sentiment.vader import SentimentIntensityAnalyzer

warnings.filterwarnings("ignore")
warnings.filterwarnings(action="ignore", category=UserWarning, module="gensim")
warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=FutureWarning)
nltk.download("brown")
nltk.download("punkt")
nltk.download("vader_lexicon")
nlp = en_core_web_sm.load()


# URL Generation, Text Cleaning and Processing, Google Search etc. Functions
def generate_URLs(
    query, region="UnitedStates", no_of_urls=1, recency="0", input_type=""
):
    recency_output = recency
    typ = input_type
    input_no_of_urls = no_of_urls
    count_of_urls = 0
    if not (
        (recency == "")
        or (recency == "one_day")
        or (recency == "one_hour")
        or (recency == "one_month")
        or (recency == "one_week")
        or (recency == "one_year")
    ):
        print("Enter valid timeline/recency - one_day, one_hour or one_month")
    elif recency == "one_hour":
        recency = "qdr:h"
    elif recency == "one_day":
        recency = "qdr:d"
    elif recency == "one_month":
        recency = "qdr:m"
    elif recency == "one_week":
        recency = "qdr:w"
    elif recency == "one_year":
        recency = "qdr:y"

    # there are other types like shopping, books, applications
    if not (
        (typ == "all")
        or (typ == "")
        or (typ == "news")
        or (typ == "images")
        or (typ == "videos")
    ):
        print("Enter valid search type - all, news, images, videos")
    elif typ == "all":
        typ = ""
    elif typ == "news":
        typ = "nws"
    elif typ == "images":
        typ = "isch"
    elif typ == "videos":
        typ = "vid"

    url_list = []
    checking_list = []
    print("Extracting the list of URLs..")
    for i in search(query, num_results=10, lang="en"):
        try:
            index = i.index("/", 10)
            mini_url = i[0:index]
            if (
                (mini_url in checking_list)
                or (i.find(".php") >= 0)
                or (i.find("/php") >= 0)
                or (i.find(".pdf") >= 0)
                or (i.find("/pdf") >= 0)
                or (i.find(".jpg") >= 0)
                or (i.find(".jpeg") >= 0)
                or (i.find("/jpg") >= 0)
                or (i.find("/jpeg") >= 0)
                or (i.find(".gif") >= 0)
                or (i.find("/gif") >= 0)
                or (i.find("/document") >= 0)
                or (i.find("www.youtube.com") >= 0)
                or (i.find(".txt") >= 0)
                or (i.find("/txt") >= 0)
                or (i.find("/docx") >= 0)
                or (i.find(".docx") >= 0)
                or (i.find(".xlsx") >= 0)
                or (i.find("/xlsx") >= 0)
                or (i.find("/ppt") >= 0)
                or (i.find(".ppt") >= 0)
            ):
                continue
            else:
                checking_list.append(mini_url)
                url_list.append(i)
                count_of_urls += 1
                if count_of_urls >= input_no_of_urls:
                    break
            no_of_urls += int(no_of_urls / 2)
        except Exception as e:
            print(e)
            pass

    print("List of URLs for " + str(query) + " retrieved.")
    return url_list, recency_output


def remove_special_characters(data):
    bad_chars = [
        "`",
        "~",
        "@",
        "#",
        "^",
        "*",
        "<",
        ">",
        "{",
        "}",
        "+",
        "=",
        "\\",
        "|",
        "$",
        "%",
        "(",
        ")",
        "[",
        "]",
        "_",
        "-",
    ]
    for i in bad_chars:
        data = data.replace(i, " ")
    return remove_multiple_whitespaces(data)


def remove_non_alphanum(data):
    return remove_multiple_whitespaces(strip_non_alphanum(data))


def remove_num(data):
    return re.sub(r"\d+", "", data)


def remove_single_letters(data):
    return re.sub(r"\b[a-zA-Z]\b", "", data)


def remove_multiple_whitespaces(data):
    try:
        return strip_multiple_whitespaces(data)
    except:
        return data


def table_pandas(page):
    return pd.read_html(page.text)


def table_html(page):
    from html_table_parser import HTMLTableParser

    try:
        xhtml = page.content.decode("utf-8")
        html_table = HTMLTableParser()
        html_table.feed(xhtml)
        return html_table.tables
    except:
        return []


def make_df_html(table):
    result = []
    for i in range(1, len(table)):
        result.append(table[i])
    result_df = pd.DataFrame(result)
    return result_df


def extract_table_list(page):
    try:
        table_list = table_pandas(page)
        table_flag = "pd"
    except:
        table_list = table_html(page)
        table_flag = "html"
        if len(table_list) == 0:
            put_tables = "No tables for this URL"
            table_flag = "no"

    return table_list, table_flag


def text_trafilatura(url):
    downloaded = trafilatura.fetch_url(url)
    return trafilatura.extract(downloaded)


def text_newspaper(url):
    article = Article(url)
    article.download()
    article.parse()
    article.nlp()
    title = article.title
    keywords = str(article.keywords)
    try:
        date1 = article.publish_date.strftime("%m/%d/%Y - %H:%M:%S")
    except:
        date1 = "N.A"
    text = article.text
    return title, keywords, date1, text


def extract_text_data(url, recency="old"):
    try:
        put_title, put_keywords, put_publishdate, put_data_np = text_newspaper(url)
        put_data_tr = text_trafilatura(url)
        if put_publishdate == "N.A":
            put_publishdate = recency + " old"
    except:
        put_data_np = "Unable to retrieve information"
        put_data_tr = "Unable to retrieve information"
        put_title = "Unable to retrieve information"
        put_keywords = "Unable to retrieve information"
        put_publishdate = "Unable to retrieve information"

    return put_title, put_keywords, put_publishdate, put_data_np, put_data_tr


def text_beautifulsoup3(page):
    soup = BeautifulSoup(page.content, "html.parser")
    return soup.get_text()


def get_data_from_URLs(query, url_list, recency, tables=False):
    main_df = pd.DataFrame(
        columns=[
            "URL",
            "Title",
            "Keywords",
            "Publish Date/ Recency",
            "Data",
            "Data with bs4",
            "Tables(if any)",
        ]
    )
    url_count = 1
    folder_name = query.replace(" ", "")
    main_file_name = "Topics/" + folder_name + "/Data/" + folder_name + ".xlsx"
    if os.path.exists(main_file_name):
        ExcelWorkbook = load_workbook(main_file_name)
        writer = pd.ExcelWriter(main_file_name, engine="openpyxl")
        writer.book = ExcelWorkbook
    else:
        writer = pd.ExcelWriter(main_file_name, engine="xlsxwriter")
    print("Extracting data from URLs..")
    for url in url_list:
        try:
            put_url = url
            try:
                page = requests.get(url, timeout=10)
            except Exception as e:
                print("Error with URL - ", url)
                print(e)
                continue
            #   FOR TEXT DATA :-
            (
                put_title,
                put_keywords,
                put_publishdate,
                put_data_np,
                put_data_tr,
            ) = extract_text_data(url, recency)
            if not put_data_tr:
                put_data_tr = "No Data Found"

            if not put_data_np:
                put_data_np = "No Data Found"

            if len(put_data_np) > len(put_data_tr):
                put_data = put_data_np
            else:
                put_data = put_data_tr
            put_data2 = text_beautifulsoup3(page)

            #   FOR TABULAR DATA :-
            if tables:
                table_list, table_flag = extract_table_list(page)

                put_tables = "No tables"
                if table_flag != "no":
                    put_tables = "Stored in: "
                    table_count = 1
                    for table in table_list:
                        name_of_sheet = (
                            "9_table" + str(table_count) + "_url" + str(url_count)
                        )
                        put_tables += name_of_sheet + ";  "
                        if table_flag == "pd":
                            df = pd.DataFrame(table)
                        elif table_flag == "html":
                            df = make_df_html(table)
                        # colname_todrop = df.columns[0]
                        # df.drop(columns=[colname_todrop],inplace=True)
                        df.to_excel(writer, index=False, sheet_name=name_of_sheet)
                        table_count += 1

            else:
                put_tables = "Tables Disabled"

            put_data = remove_multiple_whitespaces(put_data)
            put_data2 = remove_multiple_whitespaces(put_data2)
            main_df.loc[len(main_df)] = [
                put_url,
                put_title,
                put_keywords,
                put_publishdate,
                put_data,
                put_data2,
                put_tables,
            ]
            url_count += 1
        except Exception as e:
            print(e)
            pass

    main_sheet_name = "google_search"
    del main_df["Title"]
    del main_df["Publish Date/ Recency"]
    del main_df["Data with bs4"]
    del main_df["Tables(if any)"]
    main_df["Search Terms"] = query
    main_df.to_excel(
        writer, index=False, sheet_name=main_sheet_name, encoding="utf-8-sig"
    )
    writer.save()
    writer.close()
    return main_file_name


def get_google_webscrape(
    query,
    region="",
    no_of_urls=25,
    recency="",
    typ="all",
    word_search_list=[],
    common_word_list=[],
):
    url_list, recency = generate_URLs(query, region, no_of_urls, recency, typ)
    main_file_name = get_data_from_URLs(query, url_list, recency)
    return True
